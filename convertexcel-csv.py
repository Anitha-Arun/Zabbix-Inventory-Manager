import openpyxl
import csv

def excel_to_csv(excel_file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)

    # Loop through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Open a CSV file to write the sheet content
        csv_file_name = f"{sheet_name}.csv"
        
        with open(csv_file_name, mode='w', newline="", encoding='utf-8') as file:
            writer = csv.writer(file)
            
            # Write rows from the Excel sheet to CSV
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Sheet '{sheet_name}' has been saved as {csv_file_name}")

# Example usage
excel_file = r'D:\Inventory\Updated Inventory_Device_Count_2025.xlsx'  # Provide the path to your Excel file
excel_to_csv(excel_file)
